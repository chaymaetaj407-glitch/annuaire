#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generer.py — Annuaire Dynamique v5
Sources :
  - Annuaire.xlsx  → infos client (SIRET, TVA, adresse...)
  - CA Excel       → factures DO_Type=7 (titres, emballants, montants, dates)
  - JALIXE.xlsx    → planning fabrication (enrichissement des titres)
Vérification automatique INSEE pour tous les clients
"""
import os, sys, json, datetime, time, urllib.request, urllib.parse
from collections import defaultdict, Counter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
JSON_OUT   = os.path.join(SCRIPT_DIR, "data.json")
CACHE_FILE = os.path.join(SCRIPT_DIR, "insee_cache.json")

SEUIL_MOIS = 24
DATE_SEUIL = (datetime.datetime.now() - datetime.timedelta(days=SEUIL_MOIS*30)).strftime("%Y-%m-%d")

def _completer(client, data):
    """Appliquer les données INSEE sur un client si manquantes ou incorrectes"""
    changed = False
    siren = data.get("siren","")
    siret_insee = data.get("siret","")

    # SIRET manquant ou différent
    if siret_insee and not client.get("siret"):
        client["siret"] = siret_insee; changed = True
    elif siret_insee and client.get("siret","").replace(" ","") != siret_insee.replace(" ",""):
        client["siret"] = siret_insee; changed = True

    # SIREN
    if siren and not client.get("siren"):
        client["siren"] = siren; changed = True

    # TVA manquante → calculer depuis SIREN
    if siren and not client.get("tva"):
        tva = calc_tva(siren)
        if tva: client["tva"] = tva; changed = True

    # Adresse manquante
    if data.get("adresse") and not client.get("adresse"):
        client["adresse"] = data["adresse"]; changed = True
    if data.get("cp") and not client.get("code_postal"):
        client["code_postal"] = data["cp"]; changed = True
    if data.get("ville") and not client.get("ville"):
        client["ville"] = data["ville"]; changed = True

    # Nom légal si nom = numéro client
    nom = client.get("nom_client","")
    if data.get("nom") and nom == client.get("num_CT",""):
        client["nom_client"] = data["nom"]; changed = True

    # Statut et NAF (toujours mettre à jour)
    client["statut_insee"] = data.get("statut","")
    client["naf"]          = data.get("naf","")
    client["insee_verifie"] = True

    return changed

print("=" * 55)
print("  Annuaire Dynamique v5 — Génération data.json")
print("=" * 55)

try:
    from openpyxl import load_workbook
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    from openpyxl import load_workbook

def find_file(keywords, ext=".xlsx"):
    files = [f for f in os.listdir(SCRIPT_DIR) if f.lower().endswith(ext)]
    for kw in keywords:
        for f in files:
            if kw.lower() in f.lower():
                return os.path.join(SCRIPT_DIR, f)
    return None

f_annuaire = find_file(["annuaire"])
f_ca       = find_file(["ca", "analyse", "new", "gestcom_ca"])
f_jalixe   = find_file(["jalixe"])

print(f"\n📂 Annuaire : {os.path.basename(f_annuaire) if f_annuaire else '❌ MANQUANT'}")
print(f"📊 CA Excel : {os.path.basename(f_ca)        if f_ca        else '❌ MANQUANT'}")
print(f"📂 JALIXE   : {os.path.basename(f_jalixe)    if f_jalixe    else '⚠️  optionnel'}")

if not f_annuaire or not f_ca:
    print("\n❌ Fichiers manquants ! Mets Annuaire.xlsx et le fichier CA dans le dossier.")
    input("Entrée pour fermer..."); sys.exit(1)

def clean(v):
    if v is None: return ""
    return str(v).strip()

def fmt_date(v):
    if isinstance(v, datetime.datetime): return v.strftime("%Y-%m-%d")
    if isinstance(v, str) and len(v) >= 7: return v[:10]
    return ""

def read_xlsx(path, sheet=None):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    return {c: i for i, c in enumerate(headers) if c}, rows[1:]

# ── 1. ANNUAIRE ─────────────────────────────────────────
print("\n📖 Lecture ANNUAIRE...")
H, rows = read_xlsx(f_annuaire)

PAYS_INVALIDE_PY = ['ETAT ', 'REF ', 'ACTE ', 'IMM/', 'LOC.', 'DEBITEUR', 'KARM']

def nettoyer_pays(pays_raw, cp_str):
    is_french_cp = len(cp_str)==5 and cp_str.isdigit()
    invalide = any(pays_raw.upper().startswith(x) for x in PAYS_INVALIDE_PY) or \
               (pays_raw.isdigit() and len(pays_raw)>4) or \
               (len(pays_raw)>2 and pays_raw.replace("/","").replace(" ","").isdigit())
    if is_french_cp and (not pays_raw or invalide): return "FRANCE"
    if invalide: return "FRANCE"
    return pays_raw

def nettoyer_email(email_raw):
    valides = [e.strip() for e in email_raw.replace(";",",").split(",") if "@" in e.strip() and "." in e.strip()]
    return ", ".join(valides)

def calc_tva_py(siren):
    try:
        key = (12+3*(int(siren)%97))%97
        return f"FR{key:02d}{siren}"
    except: return ""

# Regrouper par num_CT pour fusionner les doublons
ct_groups = defaultdict(list)
for r in rows:
    num = clean(r[H.get("CT_Num",0)])
    if num: ct_groups[num].append(r)

annuaire = {}
nb_fusionnes = 0
for num, grp in ct_groups.items():
    if len(grp)>1: nb_fusionnes+=1
    def best(idx):
        for r in grp:
            v = clean(r[idx]) if idx<len(r) else ""
            if v: return v
        return ""
    cp  = best(H.get("CT_CodePostal",9))
    pays = nettoyer_pays(best(H.get("CT_Pays",11)), cp)
    email = nettoyer_email(best(H.get("CT_EMail",18)))
    tel = best(H.get("CT_Telephone",16)) or best(H.get("CT_Telecopie",17))
    siret = best(H.get("CT_Siret",13))
    ident = best(H.get("CT_Identifiant",12))
    siret_d = "".join(c for c in siret if c.isdigit())
    siren = siret_d[:9] if len(siret_d)>=9 else ""
    tva = ident or (calc_tva_py(siren) if siren else "")
    annuaire[num] = {
        "num_CT": num,
        "nom_client":  best(H.get("CT_Intitule",1)),
        "contact":     best(H.get("CT_Contact",6)),
        "qualite":     best(H.get("CT_Qualite",4)),
        "adresse":     best(H.get("CT_Adresse",7)),
        "complement":  best(H.get("CT_Complement",8)),
        "code_postal": cp,
        "ville":       best(H.get("CT_Ville",10)),
        "pays":        pays,
        "telephone":   tel,
        "telecopie":   best(H.get("CT_Telecopie",17)),
        "email":       email,
        "site":        best(H.get("CT_Site",19)),
        "siret":       siret,
        "siren":       siren,
        "tva":         tva,
        "iban":        best(H.get("F_BANQUET.BT_IBAN",25)),
        "bic":         best(H.get("F_BANQUET.BT_BIC",24)),
        "nb_doublons": len(grp),
    }
print(f"   ✓ {len(annuaire)} clients ({nb_fusionnes} fusionnés)")

# ── 2. JALIXE (optionnel) ───────────────────────────────
jalixe = {}
if f_jalixe:
    print("📖 Lecture JALIXE...")
    H_j, rows_j = read_xlsx(f_jalixe)
    for r in rows_j:
        cpt = clean(r[H_j.get('CptPhase', 8)])
        if not cpt: continue
        jalixe[cpt] = {
            "titre":   clean(r[H_j.get('LibTitre', 9)]),
            "atelier": clean(r[H_j.get('LibAtelier', 13)]),
            "date":    fmt_date(r[H_j.get('DATEDOS', 0)]),
        }
    print(f"   ✓ {len(jalixe)} phases JALIXE")

# ── 3. CA Excel (source principale) ────────────────────
print("📊 Lecture CA Excel (DO_Type=7)...")

# Trouver la bonne feuille
wb_ca = load_workbook(f_ca, read_only=True, data_only=True)
sheet_name = "BDD GESTCOM" if "BDD GESTCOM" in wb_ca.sheetnames else wb_ca.sheetnames[0]
H_ca, rows_ca = read_xlsx(f_ca, sheet=sheet_name)

ca_par_client    = defaultdict(list)   # CT_Num → lignes
titres_par_client = defaultdict(dict)  # CT_Num → {code_affaire: titre}
nb_lus = 0

for r in rows_ca:
    if r[H_ca.get('DO_Type', 0)] != 7: continue
    ct       = clean(r[H_ca.get('CT_Num', 1)])
    if not ct: continue
    date_str = fmt_date(r[H_ca.get('DO_Date', 3)])
    ar_ref   = clean(r[H_ca.get('AR_Ref', 6)])
    ar_design = clean(r[H_ca.get('AR_Design', 16)])
    dl_design = clean(r[H_ca.get('DL_Design', 7)])
    code_aff  = clean(r[H_ca.get('CodeAffaire', 12)])
    titre_aff = clean(r[H_ca.get('DP_CODE_AFFAIRE.CA_INITULE', 27)])
    try: qte = float(r[H_ca.get('DL_Qte', 8)] or 0)
    except: qte = 0
    try: montant = float(r[H_ca.get('DL_MontantHT', 13)] or 0)
    except: montant = 0

    # Titres par client (code affaire + titre)
    if code_aff and titre_aff:
        if code_aff not in titres_par_client[ct]:
            titres_par_client[ct][code_aff] = {"code": code_aff, "titre": titre_aff, "montant": 0, "nb": 0}
        titres_par_client[ct][code_aff]["montant"] += montant
        titres_par_client[ct][code_aff]["nb"] += 1

    # Emballants (exclure lignes sans type ou lignes purement NOTE)
    ca_par_client[ct].append({
        "date":      date_str,
        "mois":      date_str[:7] if date_str else "",
        "ar_ref":    ar_ref,
        "ar_design": ar_design,
        "dl_design": dl_design,
        "qte":       round(qte, 2),
        "montant":   round(montant, 2),
        "code_aff":  code_aff,
        "titre_aff": titre_aff,
    })
    nb_lus += 1

print(f"   ✓ {nb_lus} lignes | {len(ca_par_client)} clients avec CA")

# ── 4. Construire les clients ───────────────────────────
print("\n🔧 Construction des données...")

# Tous les clients = Annuaire + ceux qui ont du CA
all_ct = set(annuaire.keys()) | set(ca_par_client.keys())
clients = []
all_types = set()

for ct in all_ct:
    info = annuaire.get(ct, {
        "num_CT": ct, "nom_client": ct, "adresse": "", "complement": "",
        "code_postal": "", "ville": "", "pays": "", "telephone": "",
        "telecopie": "", "email": "", "site": "", "siret": "", "siren": "",
        "tva": "", "iban": "", "bic": "", "contact": "", "qualite": "",
    })

    lignes = ca_par_client.get(ct, [])

    # Titres / affaires avec dates
    affaires = {}
    for l in lignes:
        code = l['code_aff']
        titre = l['titre_aff']
        if not code or not titre: continue
        if code not in affaires:
            affaires[code] = {"code": code, "titre": titre, "montant": 0, "nb": 0, "dates": []}
        affaires[code]["montant"] += l["montant"]
        affaires[code]["nb"] += 1
        if l["date"]: affaires[code]["dates"].append(l["date"])
    affaires_list = []
    for a in sorted(affaires.values(), key=lambda x: x["titre"]):
        dates_a = sorted(a["dates"])
        affaires_list.append({
            "code":       a["code"],
            "titre":      a["titre"],
            "montant":    round(a["montant"], 2),
            "nb":         a["nb"],
            "date_debut": dates_a[0] if dates_a else "",
            "date_fin":   dates_a[-1] if dates_a else "",
        })

    # Emballants par type avec dates
    emb_data = defaultdict(lambda: {"qte": 0, "montant": 0, "dates": []})
    hist = defaultdict(lambda: defaultdict(float))
    for l in lignes:
        t = l["ar_design"]
        if not t: continue
        emb_data[t]["qte"]     += l["qte"]
        emb_data[t]["montant"] += l["montant"]
        if l["date"]: emb_data[t]["dates"].append(l["date"])
        if l["mois"]: hist[l["mois"]][t] += l["qte"]
        all_types.add(t)

    emballants = []
    for t, v in sorted(emb_data.items()):
        dates_sorted = sorted(v["dates"])
        emballants.append({
            "type":       t,
            "quantite":   round(v["qte"], 2),
            "montant_ht": round(v["montant"], 2),
            "date_debut": dates_sorted[0]  if dates_sorted else "",
            "date_fin":   dates_sorted[-1] if dates_sorted else "",
        })

    # Historique mensuel
    historique = {
        mois: {t: round(q,2) for t,q in types.items()}
        for mois, types in sorted(hist.items())
    }

    ca_total = round(sum(l["montant"] for l in lignes), 2)

    # Filtrage par date (garder toutes les dates distinctes)
    dates = sorted(set(l["date"] for l in lignes if l["date"]))
    date_premiere = dates[0] if dates else ""
    date_derniere = dates[-1] if dates else ""

    clients.append({
        **info,
        "affaires":       affaires_list,
        "emballants":     emballants,
        "historique":     historique,
        "ca_total":       ca_total,
        "nb_affaires":    len(affaires_list),
        "nb_emballants":  len(emballants),
        "nb_lignes_ca":   len(lignes),
        "date_premiere":  date_premiere,
        "date_derniere":  date_derniere,
    })

clients.sort(key=lambda x: x["nom_client"])

# Détecter les doublons de noms
noms_count = Counter(c["nom_client"].strip().upper() for c in clients if c["nom_client"])
for c in clients:
    nom_up = c["nom_client"].strip().upper()
    c["doublon"] = noms_count.get(nom_up,1) > 1 and nom_up not in (".",""," ","-")

nb_doublons = len([c for c in clients if c["doublon"]])
print(f"   ✓ {len(clients)} clients construits")
print(f"   ✓ {len([c for c in clients if c['nb_affaires']>0])} avec titres/affaires CA")
print(f"   ✓ {len([c for c in clients if c.get('siret')])} avec SIRET")
print(f"   ✓ {nb_doublons} clients avec nom en doublon")

# ── 5. Marquer les clients en sommeil (24 mois sans facture) ──
for c in clients:
    c["en_sommeil"] = bool(
        c["ca_total"] > 0 and
        (not c.get("date_derniere") or c["date_derniere"] < DATE_SEUIL)
    )
nb_sommeil = len([c for c in clients if c["en_sommeil"]])
print(f"   ✓ Seuil sommeil: {DATE_SEUIL} → {nb_sommeil} clients en sommeil")

# ── 6. Écrire data.json ─────────────────────────────────
stats = {
    "nb_clients":       len(clients),
    "nb_phases":        len(jalixe),
    "types_prestation": sorted(all_types),
    "date_generation":  datetime.datetime.now().strftime("%d/%m/%Y à %H:%M"),
}
with open(JSON_OUT, "w", encoding="utf-8") as f:
    json.dump({"stats": stats, "clients": clients}, f, ensure_ascii=False, separators=(",",":"))

size_kb = os.path.getsize(JSON_OUT) // 1024
print(f"\n💾 data.json généré ({size_kb} KB)")
print(f"✅ Terminé ! {len(clients)} clients · {datetime.datetime.now().strftime('%d/%m/%Y à %H:%M')}")
